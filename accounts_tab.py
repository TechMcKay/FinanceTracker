import openpyxl
from PySide6 import QtWidgets
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QTableWidget, QTableWidgetItem, QHeaderView, QTreeWidget
)

from ft_mainwindow import Ui_MainWindow

# Import existing excel data
transactions_spreadsheet = openpyxl.load_workbook("transactionsSpreadsheet.xlsx")
accounts_tab_sheet = transactions_spreadsheet["Sheet2"]


class AccountsTab(Ui_MainWindow):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        # Setting up QTableWidget
        # self.accounts_table = self.parent.accounts_table.findChild(QTableWidget, "accounts_table")
        # self.parent.accounts_table.setColumnCount(accounts_tab_sheet.max_column)
        # self.parent.accounts_table.setRowCount(accounts_tab_sheet.max_row - 1)
        # self.parent.accounts_table.setHorizontalHeaderLabels(
        #     ['Name of Account', 'Type of Account', 'Amount in Account'])

        # Setting up Accounts Tree(QTreeWidget).
        self.accounts_tree = self.parent.accounts_tree.findChild(QTreeWidget, "accounts_tree")
        self.parent.accounts_tree.setColumnCount(2)
        self.parent.accounts_tree.setHeaderLabels(['Accounts', 'Amount'])
        self.parent.accounts_tree.setColumnWidth(0, 300)

        # Populate accounts table with existing accounts from spreadsheet.
        # for row in range(2, accounts_tab_sheet.max_row + 1):
        #     for column in range(1, accounts_tab_sheet.max_column + 1):
        #         cell_value = accounts_tab_sheet.cell(row, column).value
        #         if column == 3:
        #             item = QTableWidgetItem(str('${:,.2f}'.format(cell_value)))  # Format to dollar format.
        #         else:
        #             item = QTableWidgetItem(str(cell_value))
        #         item.setTextAlignment(Qt.AlignCenter)
        #         self.parent.accounts_table.setItem(row - 2, column - 1, item)
        # header = self.parent.accounts_table.horizontalHeader()
        # header.setSectionResizeMode(QHeaderView.Stretch)

        # Populate the Accounts Tree with categories of accounts.
        category_of_account_list = []
        for category_of_account in accounts_tab_sheet['B']:
            if category_of_account.row != 1:
                category_of_account = str(category_of_account.value)
                category_of_account = category_of_account.capitalize()
                if category_of_account not in category_of_account_list:
                    category_of_account_list.append(category_of_account)
        items = [QtWidgets.QTreeWidgetItem([item]) for item in category_of_account_list]
        self.parent.accounts_tree.addTopLevelItems(items)
        # Populate Accounts Tree with accounts under associated categories.
        for account_type in category_of_account_list:
            account_name_list = []
            account_amount = []
            # Gathering name of accounts and amounts.
            for row in accounts_tab_sheet.iter_rows(values_only=True):
                if row[1] == account_type:
                    account_name_list.append(row[0])
                    account_amount.append(row[2])
            # Creating the QTreeWidgetItem with associated amount.
            account_category_children = []
            for i, child in enumerate(account_name_list):
                child = QtWidgets.QTreeWidgetItem([child])
                child.setText(1, str('${:,.2f}'.format(account_amount[i])))  # Format to dollar format.
                account_category_children.append(child)

            top_level_item = self.parent.accounts_tree.topLevelItem(category_of_account_list.index(account_type))
            category_with_sum = top_level_item.text(0) + f"({round((sum(account_amount)), 2)})"
            top_level_item.setText(0, category_with_sum)
            top_level_item.addChildren(account_category_children)
        # Resize first column to fit all contents.
        self.parent.accounts_tree.expandAll()
        # self.parent.accounts_tree.resizeColumnToContents(0)



