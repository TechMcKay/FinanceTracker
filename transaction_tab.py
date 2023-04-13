from ft_mainwindow import Ui_MainWindow
from data_base_module import TransactionDatabase
from emptyTransaction_Dialog import Ui_Dialog
from transactionCategoryWindow import Ui_transactionCategoryWindow
from PySide6.QtWidgets import (
    QTableView, QTableWidgetItem, QPushButton, QDateEdit,
    QComboBox, QPlainTextEdit, QLineEdit, QAbstractItemView,
    QDialog, QMainWindow
)
from PySide6.QtGui import QRegularExpressionValidator, QAction
from PySide6.QtCore import QRegularExpression, QDate, QCoreApplication, Signal
import pandas as pd
import openpyxl

# Create connect to database
finance_tracker_db = TransactionDatabase()

transaction_data = pd.read_excel('transactionsSpreadsheet.xlsx', 0)
accounts_data = pd.read_excel('transactionsSpreadsheet.xlsx', 1)

transaction_spreadsheet = openpyxl.load_workbook("transactionsSpreadsheet.xlsx")
transaction_sheet = transaction_spreadsheet["Sheet1"]
transaction_category_sheet = transaction_spreadsheet["Sheet3"]


# create Transaction Category class.
class TransactionCategoryWindow(QMainWindow, Ui_transactionCategoryWindow):
    """Transaction Category Window"""
    new_category_signal = Signal(str)

    def __init__(self, parent=None, transaction_tab=None):
        super().__init__(parent)
        self.ui = Ui_transactionCategoryWindow()
        self.ui.setupUi(self)
        self.transaction_tab = transaction_tab
        self.ui.add_new_categoryButton.clicked.connect(self.emit_new_category_signal)

    def emit_new_category_signal(self):
        new_category_text = self.ui.transaction_category_lineEdit.text()
        self.new_category_signal.emit(new_category_text)
        self.ui.transaction_category_lineEdit.clear()


class EmptyTransactionDlg(Ui_Dialog, QDialog):
    """Transaction Empty Dialog Box"""

    def __init__(self, parent=None):
        super().__init__()
        self.setupUi(self)


class TransactionTab(Ui_MainWindow):
    def __init__(self, parent):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.parent = parent

        self.transaction_table = self.parent.findChild(QTableView, "transaction_table")
        self.transaction_table.cellChanged.connect(self.transaction_table_modified)
        self.transaction_data_flag = False
        self.parent.addTransactionButton.clicked.connect(self.add_transaction_button_was_clicked)
        self.parent.filterButton.clicked.connect(self.filter_table)
        self.parent.deleteTransactionButton = self.parent.findChild(QPushButton, "deleteTransactionButton")
        self.parent.deleteTransactionButton.clicked.connect(self.delete_transaction)
        self.transaction_table.currentCellChanged.connect(self.row_selected)
        self.row = None
        self.parent.transactionDateEdit = self.parent.findChild(QDateEdit, "transactionDateEdit")
        self.parent.transactionDateEdit.setDate(QDate.currentDate())
        self.parent.amountLineEdit = self.parent.findChild(QLineEdit, "amountLineEdit")
        dollar_format = QRegularExpression(r'^-?\d{0,8}(\.\d{0,2})?$')
        validator = QRegularExpressionValidator(dollar_format)
        self.parent.amountLineEdit.setValidator(validator)
        self.parent.amountLineEdit.setMaxLength(11)
        self.parent.accountComboBox = self.parent.findChild(QComboBox, "accountComboBox")

        # Get rid of duplicates in the Type of Account Column to add to accountComboBox.
        type_of_accounts_to_set = set(accounts_data["Type of Account"].tolist())
        self.parent.accountComboBox.addItems(type_of_accounts_to_set)


        # Create transaction categories list from spreadsheet.

        # Set up category ComboBox signals and add Categories.
        self.categories_to_set = category_of_transactions
        self.parent.categoryComboBox = self.parent.findChild(QComboBox, "categoryComboBox")
        self.parent.categoryComboBox.addItems(self.categories_to_set)

        # Set up descriptionTextEdit signals.
        self.parent.descriptionTextEdit = self.parent.findChild(QPlainTextEdit, "descriptionTextEdit")
        self.parent.memoTextEdit = self.parent.findChild(QPlainTextEdit, "memoTextEdit")

        # Set up Add/Edit Transaction window button.

    # Filter Transaction Table function.
    def filter_table(self):
        pass

    # Clear filter on transaction table function.
    def clear_transaction_filter(self):
        pass

    # Delete transaction function.
    def delete_transaction(self):
        pass

    def transaction_table_modified(self, _row, _column):
        if self.transaction_data_flag:
            # Convert cell text to QTableWidgetItem.
            item_text = self.transaction_table.item(_row, _column).text()
            item = QTableWidgetItem(item_text)
            # Block signal while changing cell in QTableWidget.
            self.transaction_table.blockSignals(True)
            self.transaction_table.setItem(_row, _column, QTableWidgetItem(item.text()))
            self.transaction_table.blockSignals(False)

            # Update spreadsheet with edited cell
            # Add 2 to row(for header in spreadsheet and indexing used by openpyxl)
            # and 1 to column(indexing used by openpyxl).
            cell = transaction_sheet.cell(row=_row + 2, column=_column + 1)
            cell.value = item.text()
            transaction_spreadsheet.save("transactionsSpreadsheet.xlsx")

    # New transaction created method.
    def add_transaction_button_was_clicked(self):
        description = self.parent.descriptionTextEdit.toPlainText()
        category = self.parent.categoryComboBox.currentText()
        category = category.title()
        account = self.parent.accountComboBox.currentText()
        if self.parent.amountLineEdit.text():
            amount = "{:,.2f}".format(float(self.parent.amountLineEdit.text()))
        else:
            amount = False
        # amount = self.parent.amountLineEdit.text()
        date = self.parent.transactionDateEdit.date().toString("MM-dd-yyyy")
        memo = self.parent.memoTextEdit.toPlainText()
        # Creation of data to add list.
        data_to_add = [description, category, account, amount, date, memo]
        # Checking for non NaN data.
        if data_to_add[0] and data_to_add[3]:
            # Add data to transaction data base.
            finance_tracker_db.insert_row_of_data("transaction_database", tuple(data_to_add))
            ## add to Table

            # # Add value to transaction category spreadsheet.
            # for row in transaction_category_sheet.iter_rows():
            #     # print(row[0].value, category)
            #     if row[0].value == category:
            #         old_value = row[1].value
            #         print(amount)
            #
            # # Add new row to transaction spreadsheet.
            # transaction_sheet.append(data_to_add)
            # transaction_spreadsheet.save("transactionsSpreadsheet.xlsx")
            #
            # # Add new data to QTableWidget
            # last_row_position = self.parent.transaction_table.rowCount()
            # self.parent.transaction_table.insertRow(last_row_position)
            # for i in range(self.parent.transaction_table.columnCount()):
            #     if not data_to_add[i]:
            #         data_to_add[i] = ''
            #
            #     item = QTableWidgetItem(str(data_to_add[i]))
            #     self.parent.transaction_table.setItem(last_row_position, i, item)
            #     self.parent.transaction_table.resizeColumnsToContents()
            #     self.parent.transaction_table.horizontalHeader().setStretchLastSection(True)
            #
            #     QAbstractItemView.scrollToBottom(self.transaction_table)
            # Clear box content for next transaction.
            self.parent.descriptionTextEdit.clear()
            self.parent.memoTextEdit.clear()
            self.parent.amountLineEdit.clear()
        # Dialog window pops up if cells required to be filler are empty.
        else:
            empty_transaction_dlg = EmptyTransactionDlg(self.parent)
            empty_transaction_dlg.exec()

    # Set up transaction category window.
    def open_transaction_category_window(self):
        transaction_category_window = TransactionCategoryWindow(self.parent, self)
        transaction_category_window.new_category_signal.connect(self.add_new_transaction_category)
        transaction_category_window.show()

    def add_new_transaction_category(self, new_category_text):
        if new_category_text:
            self.parent.categoryComboBox.addItem(new_category_text)
            print(f"Added category: {new_category_text}")
        else:
            print("No category added")
